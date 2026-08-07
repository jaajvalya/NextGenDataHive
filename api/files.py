"""Upload landing-zone helpers: safe names, size-capped writes, term counts."""
from __future__ import annotations

import logging
import re
import uuid
from pathlib import Path

from fastapi import HTTPException, UploadFile

from api.settings import GLOSSARY_DIR, MAX_UPLOAD_BYTES, UPLOAD_DIR

log = logging.getLogger("datahive.files")


def ensure_upload_dir() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


def ensure_glossary_dir() -> None:
    GLOSSARY_DIR.mkdir(parents=True, exist_ok=True)


def count_glossary_terms(path: Path) -> int | None:
    """Best-effort term/row count for Excel/CSV glossary uploads."""
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            with path.open("r", encoding="utf-8", errors="ignore") as fh:
                # header + data rows
                return max(0, sum(1 for _ in fh) - 1)
        if suffix in {".xlsx", ".xls"}:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            sheet = wb["Glossary"] if "Glossary" in wb.sheetnames else wb.active
            rows = 0
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    continue
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    rows += 1
            wb.close()
            return rows
    except Exception as exc:
        log.warning("glossary term count failed for %s: %s", path.name, exc)
    return None


def safe_stored_name(original: str) -> str:
    base = Path(original or "upload").name
    safe = re.sub(r"[^\w.\- ]", "_", base).strip() or "upload"
    return f"{uuid.uuid4().hex[:12]}_{safe}"


async def write_upload_file(upload: UploadFile, dest: Path) -> int:
    size = 0
    with dest.open("wb") as out:
        while True:
            chunk = await upload.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            if size > MAX_UPLOAD_BYTES:
                out.close()
                dest.unlink(missing_ok=True)
                raise HTTPException(
                    status_code=413,
                    detail=f"File exceeds {MAX_UPLOAD_BYTES // (1024 * 1024)} MB limit.",
                )
            out.write(chunk)
    return size
